#Application to add details of applicant such as name,phonenumber,placename,bodytemperature


from openpyxl import *
from tkinter import *

wb = load_workbook('application.xlsx')

sheet = wb.active


def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Phone Number"
    sheet.cell(row=1, column=3).value = "Place Name"
    sheet.cell(row=1, column=4).value = "Body Temperature"


def focus1(event):
    PhoneNumber_field.focus_set()


def focus2(event):
    placename_field.focus_set()


def focus3(event):
    bodytemperature_field.focus_set()


def clear():
    name_field.delete(0, END)
    PhoneNumber_field.delete(0, END)
    placename_field.delete(0, END)
    bodytemperature_field.delete(0, END)


def insert():
    if (name_field.get() == "" and
            PhoneNumber_field.get() == "" and
            placename_field.get() == "" and
            bodytemperature_field.get() == ""):

        print("empty input")

    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = PhoneNumber_field.get()
        sheet.cell(row=current_row + 1, column=3).value = placename_field.get()
        sheet.cell(row=current_row + 1, column=4).value = bodytemperature_field.get()

        # save the file
        wb.save('application.xlsx')

        name_field.focus_set()

        clear()

    # Driver code


if __name__ == "__main__":
    root = Tk()

    root.configure(background='light green')

    root.title("application form")

    root.geometry("600x400")

    excel()

    heading = Label(root, text="APPLICATION",font=("times new roman",20,"bold"),fg="red" ,bg="light green")

    name = Label(root,height=4, text="NAME", bg="light green")

    PhoneNumber = Label(root,height=4, text="PHONE NUMBER", bg="light green")

    placename = Label(root,height=4,text="PLACE NAME", bg="light green")

    bodytemperature = Label(root,height=4, text="BODY TEMPERATURE ", bg="light green")

    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    PhoneNumber.grid(row=2, column=0)
    placename.grid(row=3, column=0)
    bodytemperature.grid(row=4, column=0)

    name_field = Entry(root)
    PhoneNumber_field = Entry(root)
    placename_field = Entry(root)
    bodytemperature_field = Entry(root,width=1)

    name_field.bind("<Return>", focus1)

    PhoneNumber_field.bind("<Return>", focus2)

    bodytemperature_field.bind("<Return>", focus3)

   

    name_field.grid(row=1, column=1, ipadx="100")
    PhoneNumber_field.grid(row=2, column=1, ipadx="100")
    placename_field.grid(row=3, column=1, ipadx="100")
    bodytemperature_field.grid(row=4, column=1, ipadx="100")

    excel()

    submit = Button(root,width=6,height=3, text="Submit", fg="Black",
                    bg="Red", command=insert)
    submit.grid(row=5, column=1)

    # start the GUI

    root.mainloop()

